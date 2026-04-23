import io
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException, status
from fastapi.responses import StreamingResponse
from db.connection import get_pool
from dependencies import require_admin

router = APIRouter()


# ─── Dashboard ────────────────────────────────────────────────────────────────

@router.get('/dashboard')
async def dashboard(
    cohort_id: str | None = None,
    user: dict = Depends(require_admin),
    pool=Depends(get_pool),
):
    cohort_filter = "AND u.cohort_id = $1" if cohort_id else ""
    args = [cohort_id] if cohort_id else []

    learners = await pool.fetchval(
        f"SELECT COUNT(*) FROM users u WHERE u.role = 'learner' AND u.deleted_at IS NULL {cohort_filter}",
        *args,
    )
    logins_today = await pool.fetchval(
        f"""
        SELECT COUNT(*) FROM login_events le
        JOIN users u ON u.id = le.user_id
        WHERE le.logged_at >= CURRENT_DATE {cohort_filter}
        """,
        *args,
    )
    pass_rate = await pool.fetchval(
        f"""
        SELECT ROUND(
            SUM(CASE WHEN ta.passed THEN 1 ELSE 0 END)::numeric /
            NULLIF(COUNT(*), 0) * 100, 1
        )
        FROM test_attempts ta
        JOIN users u ON u.id = ta.user_id
        WHERE ta.status = 'submitted' {cohort_filter}
        """,
        *args,
    )
    avg_study_min = await pool.fetchval(
        f"""
        SELECT ROUND(AVG(stl.duration_seconds) / 60.0, 1)
        FROM study_time_logs stl
        JOIN users u ON u.id = stl.user_id
        WHERE stl.session_date >= CURRENT_DATE - 7 {cohort_filter}
        """,
        *args,
    )

    return {
        'learner_count':    learners or 0,
        'logins_today':     logins_today or 0,
        'pass_rate':        float(pass_rate) if pass_rate else 0.0,
        'avg_study_min_7d': float(avg_study_min) if avg_study_min else 0.0,
    }


# ─── Audit ────────────────────────────────────────────────────────────────────

@router.get('/audit/logins')
async def audit_logins(
    page: int = 1,
    page_size: int = 50,
    user: dict = Depends(require_admin),
    pool=Depends(get_pool),
):
    offset = (page - 1) * page_size
    rows = await pool.fetch(
        """
        SELECT le.id, u.applicant_id, u.full_name,
               le.logged_at, le.ip_address
        FROM login_events le
        JOIN users u ON u.id = le.user_id
        ORDER BY le.logged_at DESC
        LIMIT $1 OFFSET $2
        """,
        page_size, offset,
    )
    total = await pool.fetchval('SELECT COUNT(*) FROM login_events')
    return {
        'total': total,
        'page': page,
        'page_size': page_size,
        'items': [dict(r) for r in rows],
    }


@router.get('/audit/sessions')
async def audit_sessions(
    page: int = 1,
    page_size: int = 50,
    user: dict = Depends(require_admin),
    pool=Depends(get_pool),
):
    offset = (page - 1) * page_size
    rows = await pool.fetch(
        """
        SELECT ps.id, u.applicant_id, u.full_name,
               ps.script_filter, ps.mode,
               ps.correct_count, ps.incorrect_count,
               ps.streak_max, ps.duration_seconds,
               ps.created_at, ps.completed_at
        FROM practice_sessions ps
        JOIN users u ON u.id = ps.user_id
        ORDER BY ps.created_at DESC
        LIMIT $1 OFFSET $2
        """,
        page_size, offset,
    )
    total = await pool.fetchval('SELECT COUNT(*) FROM practice_sessions')
    return {
        'total': total,
        'page': page,
        'page_size': page_size,
        'items': [dict(r) for r in rows],
    }


@router.get('/audit/tests')
async def audit_tests(
    page: int = 1,
    page_size: int = 50,
    user: dict = Depends(require_admin),
    pool=Depends(get_pool),
):
    offset = (page - 1) * page_size
    rows = await pool.fetch(
        """
        SELECT ta.id, u.applicant_id, u.full_name,
               ta.attempt_number, ta.score, ta.total_questions,
               ta.passed, ta.status, ta.submitted_at
        FROM test_attempts ta
        JOIN users u ON u.id = ta.user_id
        ORDER BY ta.submitted_at DESC NULLS LAST
        LIMIT $1 OFFSET $2
        """,
        page_size, offset,
    )
    total = await pool.fetchval('SELECT COUNT(*) FROM test_attempts')
    return {
        'total': total,
        'page': page,
        'page_size': page_size,
        'items': [dict(r) for r in rows],
    }


# ─── User Excel Export ────────────────────────────────────────────────────────

@router.get('/users/export')
async def export_users(
    user: dict = Depends(require_admin),
    pool=Depends(get_pool),
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    rows = await pool.fetch(
        """
        SELECT u.applicant_id, u.full_name, u.role,
               c.name AS cohort_name,
               u.xp, u.streak_days,
               u.created_at, u.last_practice_date,
               (SELECT COUNT(*) FROM practice_sessions ps WHERE ps.user_id = u.id AND ps.completed_at IS NOT NULL) AS sessions,
               (SELECT COUNT(*) FROM test_attempts ta WHERE ta.user_id = u.id AND ta.passed) AS tests_passed
        FROM users u
        LEFT JOIN cohorts c ON c.id = u.cohort_id
        WHERE u.deleted_at IS NULL
        ORDER BY u.applicant_id
        """
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Users'

    headers = ['Applicant ID', 'Full Name', 'Role', 'Cohort', 'XP', 'Streak Days',
               'Joined', 'Last Practice', 'Sessions', 'Tests Passed']
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='7C3AED')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col)].width = max(len(h) + 4, 14)

    for row_idx, r in enumerate(rows, 2):
        values = [
            r['applicant_id'], r['full_name'], r['role'], r['cohort_name'],
            r['xp'], r['streak_days'],
            r['created_at'].strftime('%Y-%m-%d') if r['created_at'] else '',
            str(r['last_practice_date']) if r['last_practice_date'] else '',
            r['sessions'], r['tests_passed'],
        ]
        for col, v in enumerate(values, 1):
            ws.cell(row=row_idx, column=col, value=v)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename="manabo_users.xlsx"'},
    )


# ─── User Excel Import ────────────────────────────────────────────────────────

@router.post('/users/import')
async def import_users(
    file: UploadFile = File(...),
    user: dict = Depends(require_admin),
    pool=Depends(get_pool),
):
    import openpyxl

    if not file.filename or not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='File must be .xlsx')

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail='Invalid Excel file')

    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    created = 0
    skipped = 0
    errors: list[str] = []

    for idx, row in enumerate(rows, 2):
        if not row or row[0] is None:
            continue
        applicant_id = str(row[0]).strip()
        full_name    = str(row[1]).strip() if row[1] else None
        role         = str(row[2]).strip().lower() if row[2] else 'learner'

        if not applicant_id:
            errors.append(f'Row {idx}: missing applicant_id')
            continue
        if role not in ('learner', 'admin'):
            errors.append(f'Row {idx}: invalid role "{role}", defaulting to learner')
            role = 'learner'

        try:
            result = await pool.execute(
                """
                INSERT INTO users (applicant_id, full_name, role)
                VALUES ($1, $2, $3)
                ON CONFLICT (applicant_id) DO NOTHING
                """,
                applicant_id, full_name, role,
            )
            if result == 'INSERT 0 1':
                created += 1
            else:
                skipped += 1
        except Exception as e:
            errors.append(f'Row {idx}: {e}')

    return {
        'created': created,
        'skipped': skipped,
        'errors':  errors,
    }
